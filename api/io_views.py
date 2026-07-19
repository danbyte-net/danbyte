"""Round-trip export/import API — generic over any IO-capable object type.

* ``GET  /api/io/types/``            — types the user may export/import.
* ``GET  /api/io/<slug>/fields/``    — columns + field metadata for a type.
* ``GET  /api/io/<slug>/export/``    — stream the (RBAC-scoped) rows as CSV/JSON/XLSX.
* ``POST /api/io/<slug>/import/``    — upsert rows (dry-run preview + commit).

RBAC is enforced per row: creating needs ``add``, updating needs ``change``, and
the target of an update/create must fall inside the user's row scope
(``restrict_queryset`` — constraints **and** site scope). The pretty client-side
export is unrelated; this is the editable data round-trip.
"""
from __future__ import annotations

import csv
import io as _io
import json

from django.db import transaction
from django.http import HttpResponse, StreamingHttpResponse
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from auth_api import rbac
from auth_api.object_types import is_registered, model_for

from .io import io_for, io_types
from .views import _get_active_tenant

MAX_IMPORT_ROWS = 5000
MAX_XLSX_EXPORT_ROWS = 50000


def _resolve(request, slug):
    """``(tenant, handler, model)`` or a ``Response`` error."""
    if not is_registered(slug):
        return Response({"detail": "Unknown object type."}, status=400)
    handler = io_for(slug)
    if handler is None:
        return Response({"detail": "This object type isn't importable."}, status=400)
    tenant = _get_active_tenant(request)
    if tenant is None:
        return Response({"detail": "No active tenant."}, status=403)
    return tenant, handler, model_for(slug)


def _can(request, tenant, slug, action) -> bool:
    return request.user.is_superuser or rbac.has_action(
        request.user, tenant, slug, action
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def io_types_view(request):
    tenant = _get_active_tenant(request)
    out = []
    for t in io_types():
        slug = t["slug"]
        can_export = _can(request, tenant, slug, "view")
        can_import = _can(request, tenant, slug, "add") or _can(
            request, tenant, slug, "change"
        )
        if can_export or can_import:
            out.append({**t, "can_export": can_export, "can_import": can_import})
    return Response({"object_types": out})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def io_fields_view(request, slug):
    res = _resolve(request, slug)
    if isinstance(res, Response):
        return res
    _tenant, handler, _model = res
    return Response({
        "fields": handler.field_info(),
        "columns": handler.column_names(),
        "natural_key": handler.natural_key,
    })


def _scoped_qs(request, tenant, handler, model, action):
    qs = model._default_manager.filter(tenant=tenant)
    qs = rbac.restrict_queryset(qs, request.user, tenant, handler.slug, action)
    return handler.export_queryset(qs)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def io_export_view(request, slug):
    res = _resolve(request, slug)
    if isinstance(res, Response):
        return res
    tenant, handler, model = res
    if not _can(request, tenant, slug, "view"):
        return Response({"detail": f"You can't view {slug}."}, status=403)

    # NB: not ``format`` — that's DRF's reserved content-negotiation param.
    fmt = request.query_params.get("fmt", "csv").lower()
    qs = _scoped_qs(request, tenant, handler, model, "view")
    ids = request.query_params.get("ids")
    if ids:
        qs = qs.filter(pk__in=[i for i in ids.split(",") if i.strip()])
    # Optional field filters (e.g. ipaddress export scoped to ?prefix=<id>).
    # Only narrows the already RBAC-scoped queryset — a concrete model field →
    # exact match; unknown params ignored. Can't widen access, only restrict.
    field_names = {f.name for f in model._meta.concrete_fields}
    reserved = {"fmt", "ids", "format"}
    for key, val in request.query_params.items():
        if key in reserved or key not in field_names or not val:
            continue
        try:
            qs = qs.filter(**{key: val})
        except (ValueError, TypeError):
            continue
    qs = qs.order_by("created_at")
    cols = handler.column_names()
    fname = f"{slug}.{fmt if fmt != 'xlsx' else 'xlsx'}"

    if fmt == "json":
        def gen():
            yield "["
            first = True
            for obj in qs.iterator(chunk_size=500):
                yield ("" if first else ",") + json.dumps(handler.to_row(obj))
                first = False
            yield "]"
        resp = StreamingHttpResponse(gen(), content_type="application/json")
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp

    if fmt == "xlsx":
        return _export_xlsx(qs, handler, cols, slug)

    # CSV (default), streamed.
    class _Echo:
        def write(self, value):
            return value

    writer = csv.DictWriter(_Echo(), fieldnames=cols, extrasaction="ignore")

    def gen():
        yield writer.writerow(dict(zip(cols, cols)))  # header
        for obj in qs.iterator(chunk_size=500):
            yield writer.writerow(handler.to_row(obj))

    resp = StreamingHttpResponse(gen(), content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


def _export_xlsx(qs, handler, cols, slug):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = slug[:31]
    ws.append(cols)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    n = 0
    for obj in qs.iterator(chunk_size=500):
        if n >= MAX_XLSX_EXPORT_ROWS:
            break
        row = handler.to_row(obj)
        ws.append([row.get(c, "") for c in cols])
        n += 1
    for i in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    ws.freeze_panes = "A2"
    buf = _io.BytesIO()
    wb.save(buf)
    resp = HttpResponse(
        buf.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    resp["Content-Disposition"] = f'attachment; filename="{slug}.xlsx"'
    return resp


def _parse_upload(request):
    """Rows from a multipart xlsx ``file``, a pre-parsed JSON ``rows`` array, or
    raw ``content`` + ``format`` (csv/json)."""
    upload = request.FILES.get("file")
    if upload is not None:
        from openpyxl import load_workbook

        wb = load_workbook(upload, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [(str(h).strip() if h is not None else "") for h in rows[0]]
        out = []
        for r in rows[1:]:
            d = {h: ("" if v is None else str(v)) for h, v in zip(headers, r)}
            if any(v != "" for v in d.values()):
                out.append(d)
        return out

    data = request.data or {}
    if isinstance(data.get("rows"), list):
        return [r for r in data["rows"] if isinstance(r, dict)]
    from .bulk_import import parse_rows

    return parse_rows(data.get("content", ""), data.get("format", "csv"))


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def io_import_view(request, slug):
    res = _resolve(request, slug)
    if isinstance(res, Response):
        return res
    tenant, handler, model = res
    can_add = _can(request, tenant, slug, "add")
    can_change = _can(request, tenant, slug, "change")
    if not (can_add or can_change):
        return Response({"detail": f"You can't import {slug}."}, status=403)
    dry_run = str(request.data.get("dry_run", "")).lower() in ("1", "true", "yes") \
        or request.data.get("dry_run") is True

    try:
        rows = _parse_upload(request)
    except Exception as exc:  # noqa: BLE001
        return Response({"detail": f"Couldn't parse file: {exc}"}, status=400)
    if not rows:
        return Response({"detail": "No rows found."}, status=400)
    if len(rows) > MAX_IMPORT_ROWS:
        return Response(
            {"detail": f"Too many rows (max {MAX_IMPORT_ROWS})."}, status=400
        )

    # Row scope for update-target / create-target enforcement.
    change_qs = rbac.restrict_queryset(
        model._default_manager.filter(tenant=tenant), request.user, tenant,
        slug, "change",
    )
    add_qs = rbac.restrict_queryset(
        model._default_manager.filter(tenant=tenant), request.user, tenant,
        slug, "add",
    )

    created = updated = 0
    errors, preview = [], []
    for i, row in enumerate(rows, start=1):
        try:
            with transaction.atomic():
                existing = handler.lookup(row, tenant, request.user)
                if existing is not None and not change_qs.filter(
                    pk=existing.pk
                ).exists():
                    raise PermissionRow("not permitted to update this row")
                obj, action, changes, tag_names = handler.apply(
                    existing, row, tenant, request.user
                )
                if action == "create" and not can_add:
                    raise PermissionRow("creating new rows needs 'add' permission")
                if action == "update" and not can_change:
                    raise PermissionRow("updating rows needs 'change' permission")

                if dry_run:
                    transaction.set_rollback(True)
                else:
                    handler.commit(obj, tag_names)
                    # Site-scope guard for creates: the saved row must be in the
                    # user's add-scope (mirrors the viewset write guard).
                    if action == "create" and not add_qs.filter(
                        pk=obj.pk
                    ).exists():
                        raise PermissionRow(
                            "the new row falls outside the sites you may edit"
                        )
                if action == "create":
                    created += 1
                else:
                    updated += 1
                if dry_run:
                    preview.append({
                        "row": i, "action": action,
                        "key": _row_key(handler, row), "changes": changes,
                    })
        except PermissionRow as exc:
            errors.append({"row": i, "error": str(exc), "action": "permission"})
        except Exception as exc:  # noqa: BLE001
            msgs = getattr(exc, "messages", None)
            errors.append({
                "row": i,
                "error": "; ".join(msgs) if msgs else str(exc),
            })

    return Response({
        "total": len(rows), "created": created, "updated": updated,
        "errors": errors, "dry_run": dry_run, "preview": preview,
    })


class PermissionRow(Exception):
    """A per-row RBAC rejection (rolled back, reported, batch continues)."""


def _row_key(handler, row) -> str:
    if handler.natural_key:
        return " / ".join(str(row.get(k, "")) for k in handler.natural_key)
    return str(row.get("id", ""))
